import os
import pandas as pd
import openpyxl

ColsName = [
    "EMPLOYEEID",
    "FIRSTNAME",
    "LASTNAME",
    "EMAIL",
    "PHONENUMBER",
    "HIREDATE",
    "JOBID",
    "SALARY",
    "COMMISSIONPCT",
    "MANAGERID",
    "DEPARTMENTID",
]

filePath = "C:/Shailendra/CVS/Python/data/"

fileName_master_1 = "masterdata emp.xlsx"
fileName_2 = "employees.xlsx"

try:
    book = openpyxl.load_workbook(filePath + fileName_2)
    sheetNames = book.sheetnames
    cnt = 0
    for name in sheetNames:
        cnt = cnt + 1
        print(name)
    sheet = book["employees"]
    print("sheet.max_row", sheet.max_row)
    print("sheet.max_row", sheet.max_column)

    book.close()
    df_1 = pd.read_excel(
        filePath + fileName_master_1,
        sheet_name="Master EMP",
        header=None,
        skiprows=1,
        usecols=[0],  # index of the column
        names=["NAME"],
    )

    df_2 = pd.read_excel(
        filePath + fileName_2,
        sheet_name="employees",
        header=None,
        skiprows=1,
        names=ColsName,
        # usecols=["Employee Name"]
    )
    filtered_df = df_2[df_2["FIRSTNAME"].isin(df_1["NAME"])]
    filtered_df.to_excel("Filtered_Employees.xlsx", index=False)

    print(os.getcwd())
    # print(filtered_df)
    # print(df)
    print("Done")
# print(df_1.columns)  # Shows column names
# print(df_1.head())  # Shows the first 5 rows

# print(df_2.columns)  # Shows column names
# print(df_2.head())  # Shows the first 5 rows

# firstname_column = df_2["FIRSTNAME"]
# print(firstname_column)
except Exception as e:
    print("There was an error: " + filePath + fileName_2)
    raise e
